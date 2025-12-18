import pickle
import time
import threading
import os
from openpyxl import load_workbook

# Selenium と WebDriver-Manager をインポート
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException, ElementClickInterceptedException
from selenium.webdriver.chrome.options import Options
import undetected_chromedriver as uc

# --- 定数と設定 ---
# シートの列定数
COL_SYMBOL = 0
COL_RESEARCH_TEXT = 24

# 投資判断を実施するティッカーシンボル群
analysis_tickers = """
COMP
HBNC
FSTR
GRWG
RDNW
"""

# Excelファイルのパスを指定
excel_file_path = 'G:\\マイドライブ\\Investment\\InvestmentList.xlsx'

# Excelが大きくなりすぎるとなぜか読み込めないので、ローカルに行数を減らしたExcelを用意する必要あり
#excel_file_path = r"C:\Users\uchida\Downloads\InvestmentList.xlsx"

# 実行用フォルダ
runtime_folder = 'C:\\Users\\uchida\\GeminiAutomation'

# GeminiサイトのCookieファイル
cookie_file_path = f'{runtime_folder}\\google_cookies.pkl'

# --- 関数定義 ---

def setup_driver_with_cookies(cookie_file):
    """
    google-colab-seleniumを使用してドライバをセットアップし、
    保存されたCookieを読み込んでログイン状態を復元します。
    """
    # google-colab-seleniumを使用してUndetected Chromeドライバを初期化
    # Chromeのオプションを設定
    options = uc.ChromeOptions() # undetected_chromedriver の Options を使用
    #options.add_argument('--headless')  # ヘッドレスモードを有効化
    #options.add_argument('--no-sandbox')
    #options.add_argument('--disable-dev-shm-usage')

    # --- 修正・追加 --- : 言語を日本語に設定
    options.add_argument("--lang=ja-JP")
    options.add_experimental_option('prefs', {'intl.accept_languages': 'ja,en-US,en'})
    options.add_argument('--window-size=1920,1080')

    # google-colab-seleniumを使用してChromeドライバを初期化
    driver = uc.Chrome(options=options)

    # Cookieを適用するために、まずドメインにアクセスする必要がある
    driver.get("https://gemini.google.com/?hl=ja")

    try:
        # 保存されたCookieを読み込む
        with open(cookie_file, 'rb') as f:
            cookies = pickle.load(f)

        # Cookieをドライバセッションに追加
        for cookie in cookies:
            # 'expiry'キーが浮動小数点数の場合、整数に変換
            if 'expiry' in cookie and isinstance(cookie['expiry'], float):
                cookie['expiry'] = int(cookie['expiry'])
            driver.add_cookie(cookie)

        print("Cookieの読み込みに成功しました。")
        # --- 修正 ---: refresh()の代わりに目的のページに直接移動
        driver.get("https://gemini.google.com/?hl=ja")
        time.sleep(5) # ページの読み込みを待つ

    except FileNotFoundError:
        print(f"'{cookie_file}' が見つかりません。手動ログイン用のスクリプトを実行して作成し、Colabにアップロードしてください。")
        driver.quit()
        return None
    except Exception as e:
        print(f"Cookieの読み込み中にエラーが発生しました: {e}")
        driver.quit()
        return None

    return driver

def automate_deep_research(driver:  uc.Chrome, ticker: str, prompt: str, thread_index: int):
    """
    指定されたプロンプトでGemini Deep Researchを自動化し、結果を返します。
    この関数は外部で生成されたWebDriverインスタンスを受け取ります。
    """
    try:
        # 各ティッカーの処理前に、Geminiのトップページに移動して状態をリセット
        driver.get("https://gemini.google.com/?hl=ja")
        driver.save_screenshot(f'{runtime_folder}\\初期画面{thread_index}_{ticker}.png')

        # Cookieの有効性チェック (ログインボタンの有無で判断)
        try:
            login_button_xpath = "//a[contains(@aria-label, 'ログイン')]"
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, login_button_xpath)))
            driver.save_screenshot(f'{runtime_folder}\\Cookie期限切れ{thread_index}_{ticker}.png')
            print(f"[Thread{thread_index} {ticker}] Cookieが期限切れか、ログインしていません。")
            return False
        except TimeoutException:
            print(f"[Thread{thread_index} {ticker}] ログイン状態を確認しました。Deep Researchを開始します。")

        # "ツール" ボタンをクリック
        retries = 3
        for attempt in range(retries):
            try:
                deep_research_button_xpath = "//button[contains(., 'ツール')]"
                WebDriverWait(driver, 60).until(
                    EC.element_to_be_clickable((By.XPATH, deep_research_button_xpath))
                ).click()
                print(f"[Thread{thread_index} {ticker}] ツールボタンをクリックしました。")
                break
            except (ElementClickInterceptedException) as e:
                print(f"[Thread{thread_index} {ticker}] プロンプト入力に失敗 ({type(e).__name__})。機能紹介ダイアログが表示されている可能性があります。リトライします... ({attempt + 1}/{retries})")
                driver.save_screenshot(f'{runtime_folder}\\ダイアログ表示によるプロンプト入力失敗_{thread_index}_{ticker}.png')
                dont_use_button_xpath = "//button[contains(., '利用しない')]"
                WebDriverWait(driver, 60).until(
                    EC.element_to_be_clickable((By.XPATH, dont_use_button_xpath))
                ).click()
                time.sleep(5)
                if attempt == retries - 1:
                    driver.save_screenshot(f'{runtime_folder}\\プロンプト入力失敗_{thread_index}_{ticker}.png')
                    return False
            except (ElementClickInterceptedException) as e:
                print(f"[Thread{thread_index} {ticker}] プロンプト入力に失敗 ({type(e).__name__})。リトライします... ({attempt + 1}/{retries})")
                time.sleep(3)
                if attempt == retries - 1:
                    driver.save_screenshot(f'{runtime_folder}\\プロンプト入力失敗_{thread_index}_{ticker}.png')
                    return False

        # "Deep Research" ボタンをクリック
        retries = 3
        for attempt in range(retries):
            try:
                deep_research_button_xpath = "//button[contains(., 'Deep Research')]"
                WebDriverWait(driver, 60).until(
                    EC.element_to_be_clickable((By.XPATH, deep_research_button_xpath))
                ).click()
                print(f"[Thread{thread_index} {ticker}] Deep Researchボタンをクリックしました。")
                break
            except (ElementClickInterceptedException) as e:
                print(f"[Thread{thread_index} {ticker}] プロンプト入力に失敗 ({type(e).__name__})。機能紹介ダイアログが表示されている可能性があります。リトライします... ({attempt + 1}/{retries})")
                driver.save_screenshot(f'{runtime_folder}\\ダイアログ表示によるプロンプト入力失敗_{thread_index}_{ticker}.png')
                dont_use_button_xpath = "//button[contains(., '利用しない')]"
                WebDriverWait(driver, 60).until(
                    EC.element_to_be_clickable((By.XPATH, dont_use_button_xpath))
                ).click()
                time.sleep(5)
                if attempt == retries - 1:
                    driver.save_screenshot(f'{runtime_folder}\\プロンプト入力失敗_{thread_index}_{ticker}.png')
                    return False
            except (ElementClickInterceptedException) as e:
                print(f"[Thread{thread_index} {ticker}] プロンプト入力に失敗 ({type(e).__name__})。リトライします... ({attempt + 1}/{retries})")
                time.sleep(3)
                if attempt == retries - 1:
                    driver.save_screenshot(f'{runtime_folder}\\プロンプト入力失敗_{thread_index}_{ticker}.png')
                    return False

        # プロンプト入力エリアの処理 (リトライ機構付き)
        retries = 3
        for attempt in range(retries):
            try:
                prompt_textarea_xpath = "//div[contains(@class, 'new-input-ui')]"
                prompt_textarea = WebDriverWait(driver, 30).until(
                    EC.visibility_of_element_located((By.XPATH, prompt_textarea_xpath))
                )
                driver.execute_script("arguments[0].innerText = arguments[1]", prompt_textarea, prompt)
                print(f"[Thread{thread_index} {ticker}] プロンプトを入力しました。")
                break
            except (StaleElementReferenceException, ElementClickInterceptedException) as e:
                print(f"[Thread{thread_index} {ticker}] プロンプト入力に失敗 ({type(e).__name__})。リトライします... ({attempt + 1}/{retries})")
                time.sleep(3)
                if attempt == retries - 1:
                    driver.save_screenshot(f'{runtime_folder}\\プロンプト入力失敗_{thread_index}_{ticker}.png')
                    return False

        # 送信ボタンの処理 (リトライ機構付き)
        for attempt in range(retries):
            try:
                send_button_xpath = "//button[contains(@aria-label, 'プロンプトを送信')]"
                WebDriverWait(driver, 60).until(
                    EC.element_to_be_clickable((By.XPATH, send_button_xpath))
                ).click()
                print(f"[Thread{thread_index} {ticker}] 調査を依頼しました。")
                break
            except (StaleElementReferenceException, ElementClickInterceptedException) as e:
                print(f"[Thread{thread_index} {ticker}] 送信ボタンクリックに失敗 ({type(e).__name__})。リトライします... ({attempt + 1}/{retries})")
                time.sleep(3)
                if attempt == retries - 1:
                    driver.save_screenshot(f'{runtime_folder}\\プロンプト送信失敗_{thread_index}_{ticker}.png')
                    return False
        
        time.sleep(5)

        # "リサーチを開始" ボタンのクリック (リトライ機構付き)
        print(f"[Thread{thread_index} {ticker}] 「リサーチを開始」ボタンのクリックを試みます...")
        start_research_button_xpath = "//button[contains(., 'リサーチを開始')]"
        for attempt in range(retries):
            try:
                start_research_button = WebDriverWait(driver, 600).until(
                    EC.element_to_be_clickable((By.XPATH, start_research_button_xpath))
                )
                driver.execute_script("arguments[0].click();", start_research_button)
                print(f"[Thread{thread_index} {ticker}] リサーチを開始しました。完了まで最大20分程度お待ちください...")
                break
            except (StaleElementReferenceException, ElementClickInterceptedException, TimeoutException) as e:
                print(f"[Thread{thread_index} {ticker}] リサーチ開始ボタンのクリックに失敗 ({type(e).__name__})。リトライします... ({attempt + 1}/{retries})")
                time.sleep(3)
                if attempt == retries - 1:
                    driver.save_screenshot(f'{runtime_folder}\\リサーチ開始失敗_{thread_index}_{ticker}.png')
                    return False
        
        time.sleep(5)
        driver.save_screenshot(f'{runtime_folder}\\リサーチ開始{thread_index}_{ticker}.png')

        # レポート生成の完了を待機 (ポーリング処理)
        view_report_button_xpath = "//button[contains(., 'エクスポート')]"
        googledocs_export_button_xpath = "//button[contains(., 'Google ドキュメントにエクスポート')]"
        research_wait_minutes = 20
        wait_after_reload = 10
        print(f"[Thread{thread_index} {ticker}] レポート生成を待機します... (最大{research_wait_minutes}分)")
        for minute in range(research_wait_minutes):
            try:
                export_button = WebDriverWait(driver, 60 - wait_after_reload).until(
                    EC.element_to_be_clickable((By.XPATH, view_report_button_xpath))
                )
                driver.execute_script("arguments[0].click();", export_button)
                print(f"[Thread{thread_index} {ticker}] 調査レポートが生成されました。")

                googledoc_button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, googledocs_export_button_xpath))
                )
                driver.execute_script("arguments[0].click();", googledoc_button)
                time.sleep(10)
                print(f"[Thread{thread_index} {ticker}] Google Documentが生成されました。")
                driver.save_screenshot(f'{runtime_folder}\\リサーチ完了{thread_index}_{ticker}.png')
                return True
            except TimeoutException:
                print(f"[Thread{thread_index} {ticker}] {minute + 1} 分経過... 画面をリフレッシュします。")
                driver.save_screenshot(f'{runtime_folder}\\リサーチ中_{minute+1}m_{thread_index}_{ticker}.png')
                driver.refresh()
                time.sleep(wait_after_reload)
            except StaleElementReferenceException:
                print(f"[Thread{thread_index} {ticker}] ページの更新を検知しました。次の待機サイクルで再試行します。")

        print(f"[Thread{thread_index} {ticker}] {research_wait_minutes}分以内に調査が完了しませんでした。")
        driver.save_screenshot(f'{runtime_folder}\\リサーチタイムアウト{thread_index}_{ticker}.png')
        return False

    except Exception as e:
        print(f"[Thread{thread_index} {ticker}] 予期せぬエラーが発生しました: {e}")
        driver.save_screenshot(f'{runtime_folder}\\error_{thread_index}_{ticker}.png')
        return False


def get_research_text(sheet, symbol):
    """Excelシートから指定されたシンボルのリサーチテキストを取得します。"""
    for row in sheet.iter_rows(min_row=2):
        if row[COL_SYMBOL].value == symbol:
#            print(symbol, "Found ", row[COL_RESEARCH_TEXT].value, row[COL_RESEARCH_TEXT-1].value, row[COL_RESEARCH_TEXT+1].value)
            return row[COL_RESEARCH_TEXT].value
    return None

def automate_deep_research_paralell(cookie_file, sheet, tickers, thread_index):
    """
    スレッド内でWebDriverを管理し、複数のティッカーのDeep Researchを連続して実行します。
    """
    driver = None
    try:
        driver = setup_driver_with_cookies(cookie_file)

        driver.get("https://gemini.google.com/?hl=ja") # Cookieを適用するドメインにアクセス

        """
        with open(cookie_file, 'rb') as f:
            cookies = pickle.load(f)
        for cookie in cookies:
            if 'expiry' in cookie and isinstance(cookie['expiry'], float):
                cookie['expiry'] = int(cookie['expiry'])
            driver.add_cookie(cookie)
        print(f"[Thread{thread_index}] Cookieの読み込みに成功しました。")
        # Cookie適用後、ページを再読み込み
        driver.get("https://gemini.google.com/?hl=ja")
        """
        time.sleep(5)

        # --- 各ティッカーの処理 ---
        for ticker in tickers:
            print()
            print(f"[Thread{thread_index}] 分析を開始します: {ticker}")
            research_text = get_research_text(sheet, ticker)
            
            if research_text:
                success = automate_deep_research(driver, ticker, research_text, thread_index)
                if success:
                    print(f"[Thread{thread_index} {ticker}] 解析に成功しました")
                else:
                    print(f"[Thread{thread_index} {ticker}] 解析に失敗しました")
            else:
                print(f"[Thread{thread_index} {ticker}] リサーチ用のテキストが見つかりませんでした。")

    except FileNotFoundError:
        print(f"[Thread{thread_index}] Cookieファイル '{cookie_file}' が見つかりません。")
    except Exception as e:
        print(f"[Thread{thread_index}] WebDriverのセットアップまたは実行中にエラーが発生しました: {e}")
    finally:
        if driver:
            driver.quit()
        print(f"[Thread{thread_index}] すべての解析が終了し、ブラウザを閉じました。")


# --- メイン処理 ---
if analysis_tickers:
    print("\n指定されたティッカーの分析を開始します...")

    try:
        workbook = load_workbook(excel_file_path, data_only=True)
        sheet = workbook.active
        print(f"行数: {sheet.max_row}")
    except FileNotFoundError:
        print(f"Excelファイルが見つかりません: {excel_file_path}")
        # ここで処理を終了するなど、適切なエラーハンドリングを追加
    else:
        max_thread_count = 1
        tickers = [t.strip() for t in analysis_tickers.split() if t.strip()]
        
        print(f"対象ティッカー: {tickers}")

        # ティッカーを各スレッドに分割
        tickers_for_threads = [[] for _ in range(max_thread_count)]
        for i, ticker in enumerate(tickers):
            tickers_for_threads[i % max_thread_count].append(ticker)

        threads = []
        for thread_index in range(max_thread_count):
            if tickers_for_threads[thread_index]: # ティッカーが割り当てられているスレッドのみ起動
                print(f"[Thread {thread_index}] 担当ティッカー: {tickers_for_threads[thread_index]}")
                thread = threading.Thread(
                    target=automate_deep_research_paralell,
                    args=(cookie_file_path, sheet, tickers_for_threads[thread_index], thread_index)
                )
                thread.start()
                threads.append(thread)

        # 全てのスレッドの終了を待つ
        for thread in threads:
            thread.join()
        
        print("\n全ての分析スレッドが完了しました。")