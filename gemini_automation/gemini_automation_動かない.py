import time
import os
from openpyxl import load_workbook

# Selenium と undetected_chromedriver をインポート
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException, ElementClickInterceptedException, WebDriverException
# Optionsクラスは undetected_chromedriver でも使用できます
from selenium.webdriver.chrome.options import Options

# --- 定数と設定 ---
# シートの列定数
COL_SYMBOL = 0
COL_RESEARCH_TEXT = 24

# 投資判断を実施するティッカーシンボル群
analysis_tickers = "DOCS,ROOT,GRRR"

# Excelファイルのパスを指定
excel_file_path = 'G:\\マイドライブ\\Investment\\InvestmentList.xlsx'

# 実行用フォルダ
runtime_folder = 'C:\\Users\\uchida\\GeminiAutomation'
# 実行フォルダが存在しない場合は作成
if not os.path.exists(runtime_folder):
    os.makedirs(runtime_folder)

# --- 関数定義 ---

def get_login_cookies():
    """
    メインのChromeプロファイルを使用して一度だけブラウザを起動し、
    ログイン後のCookieを取得して返す。
    """
    print("\n--------------------------------------------------")
    print("ステップ1: ログイン情報の取得を開始します。")
    options = uc.ChromeOptions() # undetected_chromedriver の Options を使用
    try:
        user_name = os.getlogin()
        user_data_dir = os.path.join('C:', os.sep, 'Users', user_name, 'AppData', 'Local', 'Google', 'Chrome', 'User Data')
        
        # 起動を安定させるためのオプションを追加
        options.add_argument(f'--user-data-dir={user_data_dir}')
        options.add_argument(r'--profile-directory=Default')
        options.add_argument('--disable-background-networking') # バックグラウンド通信を無効化
        options.add_argument('--disable-sync') # 同期を無効化

    except Exception as e:
        print(f"プロファイルパスの取得に失敗しました: {e}")
        return None

    driver = None
    try:
        # ブラウザ起動処理そのものをtry...exceptで囲む
        print("プロファイルを使用してブラウザを起動します...")
        # ▼▼▼▼▼【重要】ここを修正 ▼▼▼▼▼
        # 安定性を向上させるため use_subprocess=True を追加
        driver = uc.Chrome(options=options, use_subprocess=True)
        # ▲▲▲▲▲【重要】ここまで修正 ▲▲▲▲▲

        driver.get("https://gemini.google.com/?hl=ja")
        
        print("ログインページにアクセスしました。ログイン状態を確認します...")
        
        # より安定した要素（aria-labelにアカウント情報が含まれるボタン）を待機
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "button[aria-label*='Google アカウント']"))
        )
        
        cookies = driver.get_cookies()
        print("ログイン情報の取得に成功しました。")
        return cookies
    except TimeoutException:
        print("ログイン状態の確認に失敗しました。タイムアウトしました。")
        print("手動でChromeを起動し、Googleにログインしているか確認してください。")
        return None
    except WebDriverException as e:
        print(f"Cookie取得用のブラウザ起動に失敗しました: {e}")
        print(">>ヒント: 事前に全てのChromeウィンドウが閉じていることを確認してください。<<")
        print(">>タスクマネージャー(Ctrl+Shift+Esc)で 'chrome.exe' が残っていないか確認することをお勧めします。<<")
        return None
    finally:
        if driver:
            driver.quit()
            print("ログイン情報取得用のブラウザを閉じました。")
        print("--------------------------------------------------\n")


def automate_deep_research(driver: uc.Chrome, ticker: str, prompt: str, index: int):
    """
    指定されたプロンプトでGemini Deep Researchを自動化し、結果を返します。
    """
    try:
        # 各ティッカーの処理前に、Geminiのトップページに移動して状態をリセット
        driver.get("https://gemini.google.com/?hl=ja")
        driver.save_screenshot(f'{runtime_folder}\\初期画面_{index}_{ticker}.png')

        # ログイン状態のチェックはメイン処理で行うため、ここでは簡略化
        print(f"[{index}: {ticker}] Deep Researchを開始します。")

        # "Deep Research" ボタンをクリック
        deep_research_button_xpath = "//button[contains(., 'Deep Research')]"
        WebDriverWait(driver, 60).until(
            EC.element_to_be_clickable((By.XPATH, deep_research_button_xpath))
        ).click()
        print(f"[{index}: {ticker}] Deep Researchボタンをクリックしました。")

        # (以降の処理は変更なし)
        # プロンプト入力エリアの処理 (リトライ機構付き)
        retries = 3
        for attempt in range(retries):
            try:
                prompt_textarea_xpath = "//div[contains(@class, 'new-input-ui')]"
                prompt_textarea = WebDriverWait(driver, 30).until(
                    EC.visibility_of_element_located((By.XPATH, prompt_textarea_xpath))
                )
                driver.execute_script("arguments[0].innerText = arguments[1]", prompt_textarea, prompt)
                print(f"[{index}: {ticker}] プロンプトを入力しました。")
                break
            except (StaleElementReferenceException, ElementClickInterceptedException) as e:
                print(f"[{index}: {ticker}] プロンプト入力に失敗 ({type(e).__name__})。リトライします... ({attempt + 1}/{retries})")
                time.sleep(3)
                if attempt == retries - 1:
                    driver.save_screenshot(f'{runtime_folder}\\プロンプト入力失敗_{index}_{ticker}.png')
                    return False

        # 送信ボタンの処理 (リトライ機構付き)
        for attempt in range(retries):
            try:
                send_button_xpath = "//button[contains(@aria-label, 'プロンプトを送信')]"
                WebDriverWait(driver, 60).until(
                    EC.element_to_be_clickable((By.XPATH, send_button_xpath))
                ).click()
                print(f"[{index}: {ticker}] 調査を依頼しました。")
                break
            except (StaleElementReferenceException, ElementClickInterceptedException) as e:
                print(f"[{index}: {ticker}] 送信ボタンクリックに失敗 ({type(e).__name__})。リトライします... ({attempt + 1}/{retries})")
                time.sleep(3)
                if attempt == retries - 1:
                    driver.save_screenshot(f'{runtime_folder}\\プロンプト送信失敗_{index}_{ticker}.png')
                    return False
        
        time.sleep(5)

        # "リサーチを開始" ボタンのクリック (リトライ機構付き)
        print(f"[{index}: {ticker}] 「リサーチを開始」ボタンのクリックを試みます...")
        start_research_button_xpath = "//button[contains(., 'リサーチを開始')]"
        for attempt in range(retries):
            try:
                start_research_button = WebDriverWait(driver, 600).until(
                    EC.element_to_be_clickable((By.XPATH, start_research_button_xpath))
                )
                driver.execute_script("arguments[0].click();", start_research_button)
                print(f"[{index}: {ticker}] リサーチを開始しました。完了まで最大20分程度お待ちください...")
                break
            except (StaleElementReferenceException, ElementClickInterceptedException, TimeoutException) as e:
                print(f"[{index}: {ticker}] リサーチ開始ボタンのクリックに失敗 ({type(e).__name__})。リトライします... ({attempt + 1}/{retries})")
                time.sleep(3)
                if attempt == retries - 1:
                    driver.save_screenshot(f'{runtime_folder}\\リサーチ開始失敗_{index}_{ticker}.png')
                    return False
        
        time.sleep(5)
        driver.save_screenshot(f'{runtime_folder}\\リサーチ開始_{index}_{ticker}.png')

        # レポート生成の完了を待機 (ポーリング処理)
        view_report_button_xpath = "//button[contains(., 'エクスポート')]"
        research_wait_minutes = 20
        wait_after_reload = 10
        print(f"[{index}: {ticker}] レポート生成を待機します... (最大{research_wait_minutes}分)")
        for minute in range(research_wait_minutes):
            try:
                WebDriverWait(driver, 60 - wait_after_reload).until(
                    EC.element_to_be_clickable((By.XPATH, view_report_button_xpath))
                )
                print(f"[{index}: {ticker}] 調査レポートが生成されました。")
                driver.save_screenshot(f'{runtime_folder}\\リサーチ完了_{index}_{ticker}.png')
                return True
            except TimeoutException:
                print(f"[{index}: {ticker}] {minute + 1} 分経過... 画面をリフレッシュします。")
                driver.save_screenshot(f'{runtime_folder}\\リサーチ中_{minute+1}m_{index}_{ticker}.png')
                driver.refresh()
                time.sleep(wait_after_reload)
            except StaleElementReferenceException:
                print(f"[{index}: {ticker}] ページの更新を検知しました。次の待機サイクルで再試行します。")

        print(f"[{index}: {ticker}] {research_wait_minutes}分以内に調査が完了しませんでした。")
        driver.save_screenshot(f'{runtime_folder}\\リサーチタイムアウト_{index}_{ticker}.png')
        return False

    except Exception as e:
        print(f"[{index}: {ticker}] 予期せぬエラーが発生しました: {e}")
        try:
            driver.save_screenshot(f'{runtime_folder}\\error_{index}_{ticker}.png')
        except:
            pass
        return False


def get_research_text(sheet, symbol):
    """Excelシートから指定されたシンボルのリサーチテキストを取得します。"""
    for row in sheet.iter_rows(min_row=2):
        if row[COL_SYMBOL].value == symbol:
            return row[COL_RESEARCH_TEXT].value
    return None

# --- メイン処理 ---
if __name__ == "__main__":
    # ユーザーにChromeを閉じるよう促す
    print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!! 重要 !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
    print("このスクリプトは、普段お使いのChromeのログイン情報を一時的に利用します。")
    print("エラーを防ぐため、現在開いている全てのChromeウィンドウを手動で閉じてください。")
    input("全てのChromeウィンドウを閉じたら、Enterキーを押して続行してください...")
    print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")

    # 1. 最初にログインCookieを取得する
    login_cookies = get_login_cookies()

    if login_cookies:
        # 2. クリーンなブラウザを起動して本処理を開始
        print("ステップ2: 自動化処理用のブラウザを起動します。")
        options = uc.ChromeOptions()
        options.add_argument("--lang=ja-JP")
        # undetected_chromedriverでは、add_experimental_optionは直接使えないことが多い
        # options.add_experimental_option('prefs', {'intl.accept_languages': 'ja,en-US,en'})
        options.add_argument('--window-size=1920,1080')
        # options.add_argument('--headless') # 必要に応じて有効化
        driver = None
        
        try:
            # ▼▼▼▼▼【重要】ここを修正 ▼▼▼▼▼
            # 安定性を向上させるため use_subprocess=True を追加
            driver = uc.Chrome(options=options, use_subprocess=True)
            # ▲▲▲▲▲【重要】ここまで修正 ▲▲▲▲▲
            
            # 3. Cookieを設定してログイン状態を復元
            driver.get("https://gemini.google.com/?hl=ja")
            for cookie in login_cookies:
                driver.add_cookie(cookie)
            driver.refresh()
            print("ログイン情報を復元しました。3秒後に処理を開始します。")
            time.sleep(3)

            # 4. Excelからティッカーリストを読み込み、順次処理
            try:
                workbook = load_workbook(excel_file_path, data_only=True)
                sheet = workbook.active
                tickers = [t.strip() for t in analysis_tickers.split(',') if t.strip()]
                print(f"分析対象のティッカー: {tickers}")

                for i, ticker in enumerate(tickers):
                    print(f"\n===== 分析開始 ({i+1}/{len(tickers)}): {ticker} =====")
                    research_text = get_research_text(sheet, ticker)
                    
                    if research_text:
                        success = automate_deep_research(driver, ticker, research_text, i)
                        if success:
                            print(f"===== 分析成功: {ticker} =====")
                        else:
                            print(f"===== 分析失敗: {ticker} =====")
                    else:
                        print(f"リサーチ用のテキストが見つかりませんでした: {ticker}")

            except FileNotFoundError:
                print(f"Excelファイルが見つかりません: {excel_file_path}")
            
        except Exception as e:
            print(f"自動化処理の実行中に予期せぬエラーが発生しました: {e}")
        finally:
            if driver:
                driver.quit()
            print("\n全ての分析が終了し、ブラウザを閉じました。")
    else:
        print("ログイン情報の取得に失敗したため、処理を中断します。")

